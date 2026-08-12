"""
extract.py — SEPTA OSINT Extraction Engine (v4)
=================================================
Architecture:
  collect.py  →  ALL fetching + scraping + LLM extraction  →  JSON
  extract.py  →  aggregate JSON fields + OPA lookup + Excel  →  XLSX
"""

import os
import json
import re
import time
from urllib.parse import urlparse
import requests
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Relevance sort order: confirmed-relevant pages processed first.
RELEVANCE_ORDER = {"relevant": 0, "uncertain": 1, "irrelevant": 2}

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

INPUT_EXCEL_USERS   = "assets/exportUsers_2026-7-13.xlsx"
INPUT_RAW_JSON      = "septa_social_links.json"
OUTPUT_EXCEL_REPORT = "inside_septa.xlsx"

PHILLY_CARTO_API_URL = "https://phl.carto.com/api/v2/sql"

# Social platforms the LLM can identify — maps platform key → Excel column name.
PLATFORM_COL = {
    "linkedin":  "LinkedIn",
    "facebook":  "Facebook",
    "instagram": "Instagram",
    "tiktok":    "TikTok",
    "twitter":   "X / Twitter",
    "github":    "GitHub",
    "youtube":   "YouTube",
    "reddit":    "Reddit",
    "other":     "Other Social Platforms",
}

# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def title_case_name(name: str) -> str:
    """Capitalises first letter of each word, lowercases the rest."""
    return " ".join(w.capitalize() for w in name.split())


# ---------------------------------------------------------------------------
# RESIDENTIAL ADDRESS LOOKUP  (Philadelphia OPA — two-pass, unchanged)
# ---------------------------------------------------------------------------

def _run_opa_query(sql: str) -> list:
    try:
        time.sleep(0.25)
        session = requests.Session()
        session.trust_env = False
        r = session.get(PHILLY_CARTO_API_URL, params={"q": sql}, timeout=8, verify=False)
        if r.status_code == 200:
            rows = r.json().get("rows", [])
            return [row.get("location", "").strip() for row in rows if row.get("location")]
    except Exception:
        pass
    return []


def fetch_residential_address(employee_name: str) -> tuple:
    """
    Two-pass OPA property lookup returning (best, alt1, alt2).
    Pass 1 — Exact:  LASTNAME FIRSTNAME [MI]
    Pass 2 — Fuzzy:  LASTNAME F%  (first-initial wildcard)
    """
    if not employee_name:
        return ("", "", "")

    clean = re.sub(r"[^\w\s]", "", employee_name).strip().upper()
    parts = clean.split()

    if len(parts) < 2:
        return ("No Match", "", "")

    first = parts[0]
    last  = parts[-1]
    mi    = parts[1][0] if len(parts) > 2 else None

    if mi:
        sql1 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first} {mi}%' LIMIT 5")
    else:
        sql1 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first}%' LIMIT 5")

    results = _run_opa_query(sql1)

    if not results:
        sql2 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first[0]}%' LIMIT 5")
        results = _run_opa_query(sql2)

    seen, unique = set(), []
    for addr in results:
        if addr and addr not in seen:
            seen.add(addr)
            unique.append(addr)

    if not unique:
        return ("No Match", "", "")

    return (
        unique[0] if len(unique) > 0 else "",
        unique[1] if len(unique) > 1 else "",
        unique[2] if len(unique) > 2 else "",
    )


# ---------------------------------------------------------------------------
# PER-EMPLOYEE PROCESSING  (pure aggregation — no extraction logic)
# ---------------------------------------------------------------------------

def process_employee(employee: dict, internal_profile: dict) -> dict:
    """
    Reads LLM-extracted entity fields from all assets for one employee,
    picks best values using relevance ordering, runs OPA lookup, and
    returns a flat output row ready for the Excel sheet.

    No regex. No pattern matching. No parsing cascades.
    All extraction happened in collect.py — this function only aggregates.
    """
    name   = employee.get("employee_name", "Unknown")
    assets = employee.get("discovered_assets", [])

    # Sort: relevant assets first — first-wins logic picks best values.
    assets = sorted(
        assets,
        key=lambda a: RELEVANCE_ORDER.get(a.get("relevance", "uncertain"), 1)
    )

    # Accumulated fields
    job_title       = ""
    education       = ""
    llm_location    = ""   # Used as fallback when OPA finds no Philly address
    all_emails      = []
    all_phones      = []
    linkedin_url    = ""
    social_profiles = {}   # platform → (handle, source_url)
    confidence      = "Low"
    has_docs        = "No"

    for asset in assets:
        entities = asset.get("entities", {})
        rel      = asset.get("relevance", "uncertain")
        url      = asset.get("url", "")

        # Job title — first hit from a relevant asset wins
        if not job_title and rel == "relevant":
            jt = entities.get("job_title", "").strip()
            if jt:
                job_title = jt

        # Education — same first-wins logic
        if not education and rel == "relevant":
            edu = entities.get("education", "").strip()
            if edu:
                education = edu

        # Location — first relevant LLM-extracted location for OPA fallback
        if not llm_location and rel == "relevant":
            loc = entities.get("location", "").strip()
            if loc:
                llm_location = loc

        # Emails — collect all unique personal addresses across all assets
        for em in entities.get("email_addresses", []):
            if (em and em not in all_emails
                    and not em.lower().endswith(("@septa.org", "@septapd.org"))):
                all_emails.append(em)

        # Phones — collect all unique numbers across all assets
        for ph in entities.get("phone_numbers", []):
            if ph and ph not in all_phones:
                # Ignore numbers cut off by Serper snippets and ensure at least 10 digits
                if "..." not in ph and len(re.sub(r'\D', '', ph)) >= 10:
                    all_phones.append(ph)

        # Social platforms — iterate the full list from this asset.
        for acct in entities.get("social_accounts", []):
            platform = acct.get("platform", "").lower().strip()
            handle   = acct.get("handle", "").strip()
            
            # Ignore placeholder or garbage handles
            if not handle or "not_found" in handle.lower() or "not found" in handle.lower() or handle.lower() == "none":
                continue
                
            # Enforce strict platform correctness based on the source URL domain.
            # If the link isn't explicitly from the platform's domain, it belongs in "other".
            url_lower = url.lower()
            if "linkedin.com" in url_lower:
                platform = "linkedin"
            elif "facebook.com" in url_lower:
                platform = "facebook"
            elif "instagram.com" in url_lower:
                platform = "instagram"
            elif "tiktok.com" in url_lower:
                platform = "tiktok"
            elif "twitter.com" in url_lower or "x.com" in url_lower:
                platform = "twitter"
            elif "github.com" in url_lower:
                platform = "github"
            elif "youtube.com" in url_lower:
                platform = "youtube"
            elif "reddit.com" in url_lower:
                platform = "reddit"
            else:
                platform = "other"

            if platform == "other":
                if "other" not in social_profiles:
                    social_profiles["other"] = []
                if not any(u == url for h, u in social_profiles["other"]):
                    social_profiles["other"].append((handle, url))
            else:
                if platform and handle and platform in PLATFORM_COL and platform not in social_profiles:
                    social_profiles[platform] = (handle, url)
                    if platform == "linkedin":
                        linkedin_url = url

        # Confidence scoring
        combined = (asset.get("full_content", "") + " " + asset.get("snippet", "")).lower()
        if rel == "relevant" and (
            "septa" in combined or "southeastern pennsylvania transportation" in combined
        ):
            confidence = "High"
        elif confidence != "High" and rel != "irrelevant" and name.lower() in combined:
            confidence = "Medium"

        if asset.get("is_doc"):
            has_docs = "Yes"

    # OPA residential address lookup (unchanged)
    addr_best, addr_alt1, addr_alt2 = fetch_residential_address(name)

    # Location hint: use LLM-extracted location when OPA finds nothing
    snippet_location = llm_location if addr_best in ("No Match", "") else ""

    # Retrieve (handle, url) for each standard platform
    def _social(platform):
        return social_profiles.get(platform, ("", ""))

    fb_h,  fb_u  = _social("facebook")
    ig_h,  ig_u  = _social("instagram")
    tk_h,  tk_u  = _social("tiktok")
    tw_h,  tw_u  = _social("twitter")
    gh_h,  gh_u  = _social("github")
    yt_h,  yt_u  = _social("youtube")
    rd_h,  rd_u  = _social("reddit")
    
    # Process 'other' platforms (can be multiple)
    other_list = social_profiles.get("other", [])
    if other_list:
        ot_display = "\n".join(u for h, u in other_list)
        ot_url = ""  # We display raw URLs in the text, so no single hyperlink
    else:
        ot_display = ""
        ot_url = ""

    display_name = title_case_name(name)

    return {
        # ── Display columns (written to Excel) ──────────────────────────────
        "Full Name":                    display_name,
        "Internal Department":          internal_profile.get("Internal Department", ""),
        "Internal Job Title":           internal_profile.get("Internal Job Title", ""),
        "Inferred Job Title":   job_title,
        # Social media: display text only — URLs stored in hidden _*_url cols
        "LinkedIn":                     display_name if linkedin_url else "",
        "Facebook":                     f"@{fb_h}"  if fb_h else "",
        "Instagram":                    f"@{ig_h}"  if ig_h else "",
        "TikTok":                       f"@{tk_h}"  if tk_h else "",
        "X / Twitter":                  f"@{tw_h}"  if tw_h else "",
        "GitHub":                       f"@{gh_h}"  if gh_h else "",
        "YouTube":                      f"@{yt_h}"  if yt_h else "",
        "Reddit":                       f"u/{rd_h}" if rd_h else "",
        "Other Social Platforms":       ot_display,
        "Phone Number(s)":              " | ".join(all_phones),
        "Personal Email(s)":            " | ".join(all_emails),
        "Education":                    education,
        "Residential Address (Best)":   addr_best,
        "Residential Address (Alt 1)":  addr_alt1,
        "Residential Address (Alt 2)":  addr_alt2,
        "Location Hint (Snippet)":      snippet_location,
        "Verification Confidence":      confidence,
        "Public Documents Found":       has_docs,
        # ── Hidden hyperlink data (excluded from displayed columns) ─────────
        "_li_url": linkedin_url,
        "_fb_url": fb_u,
        "_ig_url": ig_u,
        "_tk_url": tk_u,
        "_tw_url": tw_u,
        "_gh_url": gh_u,
        "_yt_url": yt_u,
        "_rd_url": rd_u,
        "_ot_url": ot_url,
    }


# ---------------------------------------------------------------------------
# EXCEL STYLING  (unchanged)
# ---------------------------------------------------------------------------

HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
HYPERLINK_FONT = Font(color="0563C1", underline="single", size=10)
NORMAL_FONT    = Font(size=10)


def style_sheet(ws) -> None:
    """Dark-blue header, auto column widths, frozen top row."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header = col_cells[0]
        header.fill      = HEADER_FILL
        header.font      = HEADER_FONT
        header.alignment = HEADER_ALIGN
        
        # Apply text wrapping and top vertical alignment to all data cells
        data_align = Alignment(wrap_text=True, vertical="top")
        for cell in col_cells[1:]:
            cell.alignment = data_align

        max_content = max(
            (len(str(c.value or "").split("\n")[0]) for c in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_content + 3, 14), 52
        )

    ws.freeze_panes = "A2"


def apply_hyperlinks(ws, df_sorted: pd.DataFrame, col_positions: dict) -> None:
    """
    Post-processes the worksheet to convert display-text cells into
    clickable hyperlinks using the _*_url hidden columns from df_sorted.
    """
    link_map = [
        ("LinkedIn",               "_li_url"),
        ("Facebook",               "_fb_url"),
        ("Instagram",              "_ig_url"),
        ("TikTok",                 "_tk_url"),
        ("X / Twitter",            "_tw_url"),
        ("GitHub",                 "_gh_url"),
        ("YouTube",                "_yt_url"),
        ("Reddit",                 "_rd_url"),
        ("Other Social Platforms", "_ot_url"),
    ]

    for row_0idx, (_, row) in enumerate(df_sorted.iterrows()):
        excel_row = row_0idx + 2   # +1 header, +1 for 1-indexing

        for display_col, url_col in link_map:
            url     = row.get(url_col, "")
            display = row.get(display_col, "")
            if url and display and display_col in col_positions:
                cell           = ws.cell(row=excel_row, column=col_positions[display_col])
                cell.value     = display
                cell.hyperlink = str(url)
                cell.font      = HYPERLINK_FONT


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print(f"[*] Loading internal user registry from: {INPUT_EXCEL_USERS}")
    if not os.path.exists(INPUT_EXCEL_USERS):
        print(f"[!] File not found: {INPUT_EXCEL_USERS}")
        return

    try:
        df_base   = pd.read_excel(INPUT_EXCEL_USERS, engine="openpyxl")
        name_col  = df_base.columns[1]   # Column B — displayName
        dept_col  = df_base.columns[3]   # Column D — department
        title_col = df_base.columns[4]   # Column E — jobTitle
    except Exception as e:
        print(f"[!] Failed to parse spreadsheet: {e}")
        return

    user_registry: dict = {}
    for _, row in df_base.iterrows():
        emp = str(row.get(name_col, "")).strip()
        if emp and emp.lower() != "nan":
            user_registry[emp] = {
                "Internal Department": str(row.get(dept_col, "")),
                "Internal Job Title":  str(row.get(title_col, "")),
            }
    print(f"[*] Loaded {len(user_registry)} internal employee records.")

    print(f"[*] Loading OSINT JSON: {INPUT_RAW_JSON}")
    if not os.path.exists(INPUT_RAW_JSON):
        print(f"[!] File not found: {INPUT_RAW_JSON}")
        return

    with open(INPUT_RAW_JSON, "r", encoding="utf-8") as f:
        master_data = json.load(f)
    print(f"[*] Loaded {len(master_data)} employee OSINT records.")

    dossier_rows: list  = []
    reference_rows: list = []

    print("[*] Running aggregation pipeline...")
    total = len(master_data)

    for idx, employee in enumerate(master_data, 1):
        name    = employee.get("employee_name", "Unknown")
        assets  = employee.get("discovered_assets", [])
        profile = user_registry.get(name, {"Internal Department": "", "Internal Job Title": ""})

        print(f"    [{idx}/{total}] {name}")

        row = process_employee(employee, profile)
        dossier_rows.append(row)

        # Source Links Log — one row per asset
        for asset in assets:
            if asset.get("url"):
                entities = asset.get("entities", {})
                reference_rows.append({
                    "Employee Name":           name,
                    "Source URL":              asset["url"],
                    "Result Title":            asset.get("result_title", ""),
                    "Snippet":                 asset.get("snippet", ""),
                    "Full Content (Preview)":  asset.get("full_content", "")[:500],
                    "Relevance":               asset.get("relevance", ""),
                    "Social Accounts":         json.dumps(entities.get("social_accounts", [])),
                    "LLM Job Title":           entities.get("job_title", ""),
                    "Is Document":             "Yes" if asset.get("is_doc") else "No",
                    "Verification Confidence": row["Verification Confidence"],
                })

    # Build full DataFrame (includes _ hidden URL columns)
    df_full = pd.DataFrame(dossier_rows)

    # Sort by confidence then name
    df_sorted = df_full.sort_values(
        by=["Verification Confidence", "Full Name"],
        ascending=[False, True],
    ).reset_index(drop=True)

    # Display columns only (exclude _ prefixed hidden URL columns)
    display_cols = [c for c in df_sorted.columns if not c.startswith("_")]
    df_display   = df_sorted[display_cols]

    # Column name → 1-indexed Excel column position
    col_positions = {col: idx + 1 for idx, col in enumerate(display_cols)}

    df_reference = pd.DataFrame(reference_rows)

    print(f"[*] Writing report: {OUTPUT_EXCEL_REPORT}")
    try:
        with pd.ExcelWriter(OUTPUT_EXCEL_REPORT, engine="openpyxl") as writer:

            # Sheet 1 — Main dossier
            df_display.to_excel(writer, index=False, sheet_name="Inside SEPTA")

            # Sheet 2 — Raw link log
            df_reference.to_excel(writer, index=False, sheet_name="Source Links Log")

            # Style both sheets
            for sheet_name in ["Inside SEPTA", "Source Links Log"]:
                style_sheet(writer.sheets[sheet_name])

            # Apply hyperlinks to the main dossier sheet
            apply_hyperlinks(
                ws            = writer.sheets["Inside SEPTA"],
                df_sorted     = df_sorted,
                col_positions = col_positions,
            )

        print(f"[++] Report saved: {OUTPUT_EXCEL_REPORT}")

    except Exception as e:
        print(f"[!] Error writing Excel file: {e}")


if __name__ == "__main__":
    main()
