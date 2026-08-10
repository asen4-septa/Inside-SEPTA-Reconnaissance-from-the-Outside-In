"""
excel_to_json.py — Converts inside_septa.xlsx → dashboard_data.json
Run this once (or after re-running extract.py) to refresh the dashboard data.
"""
from openpyxl import load_workbook
import json, re, os

# Try the live file first, fall back to OLD
EXCEL_FILE = 'inside_septa.xlsx' if os.path.exists('inside_septa.xlsx') else 'inside_septaOLD.xlsx'
OUT_FILE   = 'dashboard_data.json'

print(f'[*] Reading: {EXCEL_FILE}')
wb  = load_workbook(EXCEL_FILE)
ws  = wb['Inside SEPTA']
wsl = wb['Source Links Log']

# ── Column headers ────────────────────────────────────────────────────────────
headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]

# ── Employees sheet ──────────────────────────────────────────────────────────
employees = []
for row in range(2, ws.max_row + 1):
    rec = {}
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        val  = str(cell.value).strip() if cell.value is not None else ''
        if val in ('nan', 'None'): val = ''
        url  = cell.hyperlink.target if cell.hyperlink else ''
        rec[header] = {'value': val, 'url': url or ''}
    # Skip totally empty rows
    name = rec.get('Full Name', {}).get('value', '')
    if name:
        employees.append(rec)

print(f'[*] Processed {len(employees)} employee rows.')

# ── Source Links sheet ───────────────────────────────────────────────────────
sl_headers = [wsl.cell(row=1, column=i).value for i in range(1, wsl.max_column + 1)]
source_links = []
for row in range(2, wsl.max_row + 1):
    r = {}
    for col_num, h in enumerate(sl_headers, start=1):
        cell = wsl.cell(row=row, column=col_num)
        r[h] = str(cell.value).strip() if cell.value is not None else ''
    emp_name = r.get('Employee Name', '')
    if emp_name and emp_name not in ('nan', 'None', ''):
        source_links.append(r)

print(f'[*] Processed {len(source_links)} source link rows.')

# ── Build source link lookup: name → [links] ─────────────────────────────────
from collections import defaultdict
link_lookup = defaultdict(list)
for sl in source_links:
    name = sl.get('Employee Name', '').strip()
    if name:
        link_lookup[name].append({
            'url':        sl.get('Source URL', ''),
            'title':      sl.get('Result Title', ''),
            'snippet':    sl.get('Snippet', ''),
            'is_doc':     sl.get('Is Document', 'No') == 'Yes',
            'confidence': sl.get('Verification Confidence', ''),
        })

# ── Assemble final JSON ───────────────────────────────────────────────────────
output = {
    'meta': {
        'total_employees': len(employees),
        'total_source_links': len(source_links),
        'source_file': EXCEL_FILE,
    },
    'employees': employees,
    'source_links_by_employee': {k: v for k, v in link_lookup.items()},
}

with open(OUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

size_mb = os.path.getsize(OUT_FILE) / 1024 / 1024
print(f'[+] Saved {OUT_FILE} ({size_mb:.2f} MB)')
