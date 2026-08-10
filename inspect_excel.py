from openpyxl import load_workbook
import json

wb = load_workbook('inside_septaOLD.xlsx')
ws = wb['Inside SEPTA']
wsl = wb['Source Links Log']

headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]

# Extract all data + hyperlinks
employees = []
for row in range(2, ws.max_row + 1):
    row_data = {}
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_num)
        val = cell.value
        hl = cell.hyperlink.target if cell.hyperlink else None
        row_data[header] = {'value': str(val) if val is not None else '', 'url': hl or ''}
    employees.append(row_data)

# Source links
sl_headers = [wsl.cell(row=1, column=i).value for i in range(1, wsl.max_column+1)]
source_links = []
for row in range(2, wsl.max_row + 1):
    r = {}
    for col_num, h in enumerate(sl_headers, start=1):
        cell = wsl.cell(row=row, column=col_num)
        r[h] = str(cell.value) if cell.value is not None else ''
    source_links.append(r)

print('Total employees:', len(employees))
print('Total source links:', len(source_links))
print()
print('=== Sample Employee ===')
e = employees[0]
for k, v in e.items():
    print(f"  {k}: value={repr(v['value'])} url={repr(v['url'])}")

print()
print('Employees with real addresses:', sum(1 for e in employees if e['Residential Address (Best)']['value'] not in ('', 'No Match', 'nan')))
print('Employees with LinkedIn url:', sum(1 for e in employees if e['LinkedIn']['url']))
print('Employees with Facebook url:', sum(1 for e in employees if e['Facebook']['url']))
print('Employees with Twitter url:', sum(1 for e in employees if e['X / Twitter']['url']))
print('Employees with Instagram url:', sum(1 for e in employees if e['Instagram']['url']))
print('Employees with phones:', sum(1 for e in employees if e['Phone Number(s)']['value']))
print('Employees with emails:', sum(1 for e in employees if e['Personal Email(s)']['value']))
