"""
extract.py — SEPTA OSINT Extraction Engine (v3)
=================================================
Changes from v2:
  (1)  Sheet renamed to "Inside SEPTA"
  (2)  Full Name displayed in proper Title Case
  (3)  Job title deduction: smarter cleaning — strips date artefacts, handles
       "Joins X As Title", "Person worked as a Title" patterns, trailing verbs
  (4)  LinkedIn cell shows employee name as a clickable hyperlink (not raw URL)
  (5)  Social media cells show @handle as a clickable hyperlink (not raw URL)
  (6)  Handle validation tightened — extended skip-list, length/charset guard
  (7)  Education: pattern-based institution extraction (not freeform segment)
  (8)  Previous Employers column removed (too noisy to be reliable)
  (9)  Risk Notes column removed
"""

import os
import json
import re
import time
from urllib.parse import urlparse
import requests
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------
INPUT_EXCEL_USERS   = "assets/exportUsers_2026-7-13.xlsx"
INPUT_RAW_JSON      = "septa_social_links.json"
OUTPUT_EXCEL_REPORT = "inside_septa.xlsx"

PHILLY_CARTO_API_URL = "https://phl.carto.com/api/v2/sql"

# ---------------------------------------------------------------------------
# REFERENCE SETS
# ---------------------------------------------------------------------------

SOCIAL_PLATFORMS = {
    "facebook.com":  "Facebook",
    "instagram.com": "Instagram",
    "tiktok.com":    "TikTok",
    "twitter.com":   "X / Twitter",
    "x.com":         "X / Twitter",
    "github.com":    "GitHub",
    "youtube.com":   "YouTube",
    "reddit.com":    "Reddit",
}

DATA_BROKERS = {
    "rocketreach", "contactout", "zoominfo", "theorg",
    "radaris", "whitepages", "spokeo", "beenverified",
    "intelius", "peoplefinder", "peoplesmart", "instantcheckmate",
    "truthfinder", "familytreenow", "fastpeoplesearch", "411", "anywho",
}

GARBAGE_VERBS = {
    "shot", "killed", "arrested", "died", "murder", "stabbed",
    "attack", "charged", "convicted", "sentenced", "felony",
    "indicted", "assault", "robbery", "injured", "hospitalized",
}

CORPORATE_KEYWORDS = [
    "manager", "analyst", "director", "officer", "lead", "specialist",
    "engineer", "coordinator", "supervisor", "operator", "chief",
    "administrator", "technician", "mechanic", "inspector", "planner",
    "agent", "representative", "consultant", "developer", "architect",
    "dispatcher", "auditor", "accountant", "detective", "sergeant",
    "captain", "lieutenant", "trainer", "instructor", "clerk", "driver",
    "foreman", "superintendent", "vice president", "president",
    "secretary", "treasurer", "commissioner", "assistant", "associate",
    "intern", "contractor", "researcher", "programmer",
    "advisor", "investigator", "examiner", "controller", "buyer",
]

# Path segments that are NOT profile handles
SKIP_SEGMENTS = {
    # Generic navigation
    "pages", "groups", "share", "watch", "r", "pub", "dir", "posts",
    "in", "search", "explore", "hashtag", "topic", "events",
    "marketplace", "gaming", "help", "profile", "login", "signup",
    "register", "settings", "notifications", "messages", "trending",
    "home", "feed", "news", "about", "contact",
    # Instagram
    "reel", "reels", "p", "tv", "stories", "highlights",
    # YouTube
    "shorts", "live", "c", "channel", "featured", "videos", "playlists",
    "watch", "results",
    # Twitter / X
    "i", "compose", "intent", "tweet", "status",
    # Facebook
    "photo", "photos", "video", "videos", "friends", "permalink", "dialog",
    # Reddit
    "user", "wiki", "comments", "submit", "new",
    # LinkedIn
    "company", "school", "jobs", "learning",
    # TikTok
    "tag", "discover",
}

# Regex patterns to match institution names cleanly
# Note: patterns require at least one qualifying word after the institution type
# to prevent bare matches like "Community College" with no school name
INSTITUTION_PATTERNS = [
    re.compile(r"\bUniversity\s+of\s+[A-Z][A-Za-z\s]{3,35}", re.IGNORECASE),
    re.compile(r"\bCollege\s+of\s+[A-Z][A-Za-z\s]{3,35}", re.IGNORECASE),
    re.compile(r"\bCommunity\s+College\s+of\s+[A-Za-z\s]{4,30}", re.IGNORECASE),
    re.compile(r"\b[A-Z][A-Za-z\s]{3,30}\s+University\b", re.IGNORECASE),
    re.compile(r"\b[A-Z][A-Za-z\s]{4,30}\s+College\b", re.IGNORECASE),
    re.compile(r"\b[A-Z][A-Za-z\s]{3,30}\s+Institute(?:\s+of\s+[A-Za-z\s]+)?\b", re.IGNORECASE),
    re.compile(r"\b[A-Z][A-Za-z\s]{3,30}\s+Academy\b", re.IGNORECASE),
    re.compile(r"\b[A-Z][A-Za-z\s]{3,30}\s+School\s+of\s+[A-Za-z\s]{4,30}", re.IGNORECASE),
]

# Location hint patterns harvested from snippets (for non-Philly fallback)
ADDRESS_HINT_PATTERNS = [
    re.compile(r"based in ([A-Z][a-zA-Z\s]+,\s*[A-Z]{2})", re.IGNORECASE),
    re.compile(r"located in ([A-Z][a-zA-Z\s]+,\s*[A-Z]{2})", re.IGNORECASE),
    re.compile(r"([A-Z][a-zA-Z\s]{2,20},\s*[A-Z]{2}\s+\d{5})", re.IGNORECASE),
    re.compile(r"lives in ([A-Z][a-zA-Z\s]+(?:,\s*[A-Z]{2})?)", re.IGNORECASE),
]


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def title_case_name(name: str) -> str:
    """Capitalises first letter of each word, lowercases the rest.
    Uses word.capitalize() per token so apostrophe chars are not
    artificially uppercased (e.g. "A'chynee" stays "A'chynee")."""
    return " ".join(w.capitalize() for w in name.split())


# ---------------------------------------------------------------------------
# RESIDENTIAL ADDRESS LOOKUP  (Philadelphia OPA — two-pass)
# ---------------------------------------------------------------------------

def _run_opa_query(sql: str) -> list:
    try:
        time.sleep(0.25)
        session = requests.Session()
        session.trust_env = False
        r = session.get(PHILLY_CARTO_API_URL, params={"q": sql}, timeout=8, verify=False)
        if r.status_code == 200:
            rows = r.json().get("rows", [])
            return [row.get("location", "").strip() for row in rows if row.get("location")]
    except Exception:
        pass
    return []


def fetch_residential_address(employee_name: str) -> tuple:
    """
    Two-pass OPA property lookup returning (best, alt1, alt2).
    Pass 1 — Exact:  LASTNAME FIRSTNAME [MI]
    Pass 2 — Fuzzy:  LASTNAME F%  (first-initial wildcard)
    """
    if not employee_name:
        return ("", "", "")

    clean = re.sub(r"[^\w\s]", "", employee_name).strip().upper()
    parts = clean.split()

    if len(parts) < 2:
        return ("No Match", "", "")

    first = parts[0]
    last  = parts[-1]
    mi    = parts[1][0] if len(parts) > 2 else None

    if mi:
        sql1 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first} {mi}%' LIMIT 5")
    else:
        sql1 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first}%' LIMIT 5")

    results = _run_opa_query(sql1)

    if not results:
        sql2 = (f"SELECT location FROM opa_properties_public "
                f"WHERE owner_1 LIKE '{last} {first[0]}%' LIMIT 5")
        results = _run_opa_query(sql2)

    seen, unique = set(), []
    for addr in results:
        if addr and addr not in seen:
            seen.add(addr)
            unique.append(addr)

    if not unique:
        return ("No Match", "", "")

    return (
        unique[0] if len(unique) > 0 else "",
        unique[1] if len(unique) > 1 else "",
        unique[2] if len(unique) > 2 else "",
    )


def extract_address_hints(assets: list) -> str:
    """Fallback: harvest city/state hints from snippets for non-Philly employees."""
    counts: dict = {}
    for asset in assets:
        text = asset.get("result_title", "") + " " + asset.get("snippet", "")
        for pattern in ADDRESS_HINT_PATTERNS:
            for m in pattern.finditer(text):
                loc = m.group(1).strip().title()
                if len(loc) < 4 or "septa" in loc.lower():
                    continue
                counts[loc] = counts.get(loc, 0) + 1
    return max(counts, key=counts.get) if counts else ""


# ---------------------------------------------------------------------------
# JOB TITLE DEDUCTION  (multi-pass, source-prioritised)
# ---------------------------------------------------------------------------

def _is_garbage_title(text: str, trusted_source: bool = False) -> bool:
    """
    Returns True if the candidate title segment is noise / garbage.
    trusted_source=True (LinkedIn /in/ profiles) relaxes the capitalization
    heuristic since LinkedIn headlines legitimately use title case.
    """
    t = text.strip().lstrip(". ").rstrip(". ")
    t = re.sub(r"^\.+\s*", "", t).strip()
    if not t or len(t) > 110:
        return True
    if re.match(r"^\d+", t):
        return True
    if re.search(r"\b\d+\+", t):
        return True
    if not trusted_source:
        words = t.split()
        if len(words) >= 5:
            cap_ratio = sum(1 for w in words if w and w[0].isupper()) / len(words)
            if cap_ratio >= 0.80:
                return True
    t_lower = t.lower()
    if any(v in t_lower for v in GARBAGE_VERBS):
        return True
    if re.search(r'\d+\+?\s+["\u201c]', t):
        return True
    if "obituary" in t_lower or "in memoriam" in t_lower:
        return True
    return False


def _clean_title_text(raw: str) -> str:
    """
    Strips all filler, date artefacts, verb-phrase preambles, and trailing
    garbage from a raw segment to produce a concise job title.
    """
    c = raw.strip().lstrip(". ").rstrip(". ")
    c = re.sub(r"^\.+\s*", "", c).strip()

    # --- Strip trailing Serper truncation ("...") ---
    c = re.sub(r"\s*\.{2,}\s*$", "", c).strip()

    # --- Strip trailing date artefacts ---
    # Matches: ." October 29, 2020" or just ", 2020" or "." 2020"
    c = re.sub(
        r'[.\s]*["\u201c\u201d]?\s*'
        r'(?:January|February|March|April|May|June|July|'
        r'August|September|October|November|December)\s+\d{1,2},?\s+\d{4}'
        r'["\u201c\u201d]?\s*$',
        "", c, flags=re.IGNORECASE,
    ).strip()
    # Trailing quoted date with period: ." 2020"
    c = re.sub(r'\.\s*["\u201c\u201d][^"]{0,40}$', "", c).strip()
    # Dangling quotes at end
    c = re.sub(r'["\u201c\u201d]\s*$', "", c).strip()

    # --- Strip numeric prefix (e.g. "7800. ") ---
    c = re.sub(r"^\d+[\.\-\s]+", "", c).strip()

    # --- Handle "Joins/Joined Company As Title" → extract "Title" ---
    m = re.match(r'^(?:joins?|joined)\s+.+?\s+as\s+(.+)$', c, re.IGNORECASE)
    if m:
        c = m.group(1).strip()

    # --- Handle "Person worked/works/is/was/served as a Title" → extract "Title" ---
    m = re.match(
        r'^.{0,30}\s+(?:worked|works|is|was|served|serves)\s+as\s+an?\s+(.+)$',
        c, re.IGNORECASE,
    )
    if m:
        c = m.group(1).strip()

    # --- Strip common leading filler phrases ---
    c = re.sub(
        r"^(?:is\s+currently\s+an?|working\s+as\s+an?|currently\s+an?|"
        r"worked\s+as\s+an?|works?\s+as\s+an?|was\s+an?\s+|served\s+as\s+an?|"
        r"serves?\s+as\s+an?|they\s+work\s+as\s+an?|a\s+hired|a\s+professional)"
        r"\s+",
        "", c, flags=re.IGNORECASE,
    ).strip()

    # --- Strip leading "a / an" determiners left over ---
    c = re.sub(r"^an?\s+", "", c, flags=re.IGNORECASE).strip()

    # --- Truncate at employer transitions ---
    c = re.split(r"\s+(?:at|for|with|from|via|brings)\s+", c,
                 maxsplit=1, flags=re.IGNORECASE)[0]

    # --- Truncate at conjunctive clauses ---
    c = re.split(r"\s+(?:and\s+(?:his|her|their)|who\s+also)\s+", c,
                 maxsplit=1, flags=re.IGNORECASE)[0]

    # --- Strip trailing hanging prepositions/conjunctions ---
    c = re.sub(
        r"\s+(?:for|at|with|in|of|from|to|by|and|or|a|an)\s*$",
        "", c, flags=re.IGNORECASE,
    ).strip()

    # --- Strip residual trailing punctuation ---
    c = c.strip(" ,.-—|\"'\u201c\u201d")

    # Hard cap
    if len(c) > 75:
        c = c[:72].rstrip() + "…"

    return c.strip()


def extract_job_title(assets: list, employee_name: str) -> str:
    """
    Three-tier extraction:
      P1 — LinkedIn /in/ profile (highest trust, relaxed garbage filter)
      P2 — Known data broker aggregators
      P3 — Generic web snippets (strict filter + 65-char cap)
    """
    candidates_p1: list = []
    candidates_p2: list = []
    candidates_p3: list = []

    for asset in assets:
        url     = asset.get("url", "")
        title_t = asset.get("result_title", "")
        snippet = asset.get("snippet", "")
        if not url:
            continue

        parsed = urlparse(url)
        domain = parsed.netloc.lower().replace("www.", "")
        path   = parsed.path
        combined = title_t + " | " + snippet

        is_linkedin_profile = "linkedin.com" in domain and "/in/" in path
        is_data_broker      = any(b in domain for b in DATA_BROKERS)

        # Remove employee name noise before scanning segments
        search_text = re.sub(re.escape(employee_name), "", combined, flags=re.IGNORECASE)
        segments    = re.split(r"\s*[|\-\u2014\u2022\xb7]\s*", search_text)

        for seg in segments:
            seg_clean = seg.strip().lstrip(". ").rstrip(". ")
            seg_clean = re.sub(r"^\.+\s*", "", seg_clean).strip()
            if len(seg_clean) < 4:
                continue
            seg_lower = seg_clean.lower()

            if not any(kw in seg_lower for kw in CORPORATE_KEYWORDS):
                continue

            if is_linkedin_profile:
                if not _is_garbage_title(seg_clean, trusted_source=True):
                    candidates_p1.append(seg_clean)
            elif is_data_broker:
                if not _is_garbage_title(seg_clean, trusted_source=False):
                    candidates_p2.append(seg_clean)
            else:
                if (not _is_garbage_title(seg_clean, trusted_source=False)
                        and len(seg_clean) <= 65):
                    candidates_p3.append(seg_clean)

    for pool in [candidates_p1, candidates_p2, candidates_p3]:
        if pool:
            cleaned = _clean_title_text(pool[0])
            # Final sanity: reject if it still begins with a plain past-tense verb
            if cleaned and not re.match(
                r'^(?:worked|joined|served|was|is|has|had|did|became)\b',
                cleaned, re.IGNORECASE,
            ):
                return cleaned

    return ""


# ---------------------------------------------------------------------------
# SOCIAL MEDIA HANDLE EXTRACTION
# ---------------------------------------------------------------------------

def extract_linkedin_url(path: str) -> str:
    """Returns a well-formed LinkedIn profile URL, or empty string."""
    if "/in/" in path:
        m = re.search(r"/in/([^/?#]+)", path)
        if m:
            return f"https://linkedin.com/in/{m.group(1)}"
    return ""


def is_valid_handle(handle: str) -> bool:
    """Returns True only for plausible social media usernames."""
    if not handle:
        return False
    if len(handle) < 3 or len(handle) > 35:
        return False
    # Must contain at least one letter (filters pure numeric post IDs)
    if not any(c.isalpha() for c in handle):
        return False
    # Only typical username chars
    if not re.match(r'^[A-Za-z0-9._\-]+$', handle):
        return False
    # Skip known generic path segments
    if handle.lower() in SKIP_SEGMENTS:
        return False
    # Filter UUID / hash-like strings (all hex, 20+ chars)
    if re.match(r'^[0-9a-fA-F]{20,}$', handle):
        return False
    return True


def extract_social_handle(domain: str, path: str, query: str) -> str:
    """Extracts a validated username from a social-media URL."""
    # Facebook numeric profile IDs via ?id=...
    if "facebook.com" in domain and "profile.php" in path:
        m = re.search(r"id=(\d+)", query)
        return m.group(1) if m else ""

    # Skip Facebook /posts/... URLs — those are page activity, not personal profiles
    if "facebook.com" in domain and "/posts/" in path:
        return ""

    m = re.search(r"^/(?:@|people/|user/|channel/|c/)?([^/?#]+)", path)
    if m:
        handle = m.group(1)
        if is_valid_handle(handle):
            return handle
    return ""


# ---------------------------------------------------------------------------
# EDUCATION EXTRACTION  (pattern-based institution names)
# ---------------------------------------------------------------------------

def extract_education(assets: list) -> str:
    """
    Searches for named academic institutions using regex patterns.
    Prioritises LinkedIn and data broker sources.
    Returns just the institution name string, not a full sentence.
    """
    candidates: list = []  # (priority, length, text)

    for asset in assets:
        url  = asset.get("url", "")
        text = asset.get("result_title", "") + " " + asset.get("snippet", "")

        parsed = urlparse(url)
        domain = parsed.netloc.lower().replace("www.", "")

        is_linkedin = "linkedin.com" in domain
        is_broker   = any(b in domain for b in DATA_BROKERS)
        priority    = 1 if is_linkedin else (2 if is_broker else 3)

        # Strip "Education:" label artefacts before scanning
        text = re.sub(r'\bEducation\s*:\s*', '', text, flags=re.IGNORECASE)

        for pattern in INSTITUTION_PATTERNS:
            for m in pattern.finditer(text):
                result = m.group(0).strip()
                if not (5 < len(result) < 80):
                    continue
                # Hard filter: reject any match containing garbage words
                r_lower = result.lower()
                if any(g in r_lower for g in [
                    "kidnapping", "renowned", "remember", "married",
                    "degree murder", "degree assault",
                ]):
                    continue
                candidates.append((priority, len(result), result))

    if not candidates:
        return ""

    # Best = lowest priority number (LinkedIn first), then shortest (cleanest)
    candidates.sort(key=lambda x: (x[0], x[1]))
    return candidates[0][2]


# ---------------------------------------------------------------------------
# CONTACT INFORMATION PARSING
# ---------------------------------------------------------------------------

def parse_contacts(combined_text: str) -> tuple:
    """Returns (phones: list, personal_emails: list).
    Excludes @septa.org / @septapd.org addresses."""
    all_emails = re.findall(
        r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b",
        combined_text,
    )
    personal_emails = list({
        e for e in all_emails
        if not e.lower().endswith("@septa.org")
        and not e.lower().endswith("@septapd.org")
    })

    raw_phones = re.findall(
        r"\b(?:\+?1[-.\s]?)?\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})\b",
        combined_text,
    )
    phones, seen = [], set()
    for g in raw_phones:
        fmt = f"({g[0]}) {g[1]}-{g[2]}"
        if fmt not in seen:
            seen.add(fmt)
            phones.append(fmt)

    return phones, personal_emails


# ---------------------------------------------------------------------------
# PER-EMPLOYEE PROCESSING
# ---------------------------------------------------------------------------

def process_employee(employee: dict, internal_profile: dict) -> dict:
    """Builds one flat output row for an employee, including hidden _url_ fields."""
    name   = employee.get("employee_name", "Unknown")
    assets = employee.get("discovered_assets", [])

    # Social media: store (handle, source_url) tuples per platform
    social_cols: dict = {
        "Facebook":     [],
        "Instagram":    [],
        "TikTok":       [],
        "X / Twitter":  [],
        "GitHub":       [],
        "YouTube":      [],
        "Reddit":       [],
        "Other Social": [],
    }

    linkedin_url = ""
    all_phones: list = []
    all_emails: list = []
    confidence  = "Low"
    has_docs    = "No"

    for asset in assets:
        url     = asset.get("url", "")
        title_t = asset.get("result_title", "")
        snippet = asset.get("snippet", "")
        is_doc  = asset.get("is_doc", False)
        if not url:
            continue

        parsed  = urlparse(url)
        domain  = parsed.netloc.lower().replace("www.", "")
        path    = parsed.path
        query   = parsed.query
        combined = title_t + " | " + snippet

        # LinkedIn profile URL (first clean /in/ hit wins)
        if "linkedin.com" in domain and not linkedin_url:
            candidate = extract_linkedin_url(path)
            if candidate:
                linkedin_url = candidate

        # Per-platform social handles
        for s_domain, s_name in SOCIAL_PLATFORMS.items():
            if s_domain in domain:
                handle = extract_social_handle(domain, path, query)
                if handle:
                    col = s_name if s_name in social_cols else "Other Social"
                    if handle not in [e[0] for e in social_cols[col]]:
                        social_cols[col].append((handle, url))
                break

        # Contact info
        phones, emails = parse_contacts(combined)
        for p in phones:
            if p not in all_phones:
                all_phones.append(p)
        for e in emails:
            if e not in all_emails:
                all_emails.append(e)

        # Escalating confidence scoring
        clow = combined.lower()
        if any(kw in clow for kw in ["septa", "southeastern pennsylvania transportation"]):
            confidence = "High"
        elif confidence != "High" and name.lower() in clow:
            confidence = "Medium"

        if is_doc:
            has_docs = "Yes"

    # Job title (multi-pass over all assets)
    deduced_title = extract_job_title(assets, name)

    # Education (pattern-based)
    education = extract_education(assets)

    # Residential address (two-pass OPA + snippet fallback)
    addr_best, addr_alt1, addr_alt2 = fetch_residential_address(name)
    snippet_location = extract_address_hints(assets) if addr_best in ("No Match", "") else ""

    # Helper: first best entry per platform
    def _first(entries):
        return entries[0] if entries else ("", "")

    fb_h,  fb_u  = _first(social_cols["Facebook"])
    ig_h,  ig_u  = _first(social_cols["Instagram"])
    tk_h,  tk_u  = _first(social_cols["TikTok"])
    tw_h,  tw_u  = _first(social_cols["X / Twitter"])
    gh_h,  gh_u  = _first(social_cols["GitHub"])
    yt_h,  yt_u  = _first(social_cols["YouTube"])
    rd_h,  rd_u  = _first(social_cols["Reddit"])
    ot_h,  ot_u  = _first(social_cols["Other Social"])

    display_name = title_case_name(name)

    return {
        # ── Display columns (written to Excel) ──────────────────────────────
        "Full Name":                    display_name,
        "Internal Department":          internal_profile.get("Internal Department", ""),
        "Internal Job Title":           internal_profile.get("Internal Job Title", ""),
        "Deduced Job Title (Online)":   deduced_title,
        # Social media: display text only — URLs stored in _*_url hidden cols
        "LinkedIn":                     display_name if linkedin_url else "",
        "Facebook":                     f"@{fb_h}"  if fb_h else "",
        "Instagram":                    f"@{ig_h}"  if ig_h else "",
        "TikTok":                       f"@{tk_h}"  if tk_h else "",
        "X / Twitter":                  f"@{tw_h}"  if tw_h else "",
        "GitHub":                       f"@{gh_h}"  if gh_h else "",
        "YouTube":                      f"@{yt_h}"  if yt_h else "",
        "Reddit":                       f"u/{rd_h}" if rd_h else "",
        "Other Social Platforms":       f"@{ot_h}"  if ot_h else "",
        "Phone Number(s)":              " | ".join(all_phones),
        "Personal Email(s)":            " | ".join(all_emails),
        "Education":                    education,
        "Residential Address (Best)":   addr_best,
        "Residential Address (Alt 1)":  addr_alt1,
        "Residential Address (Alt 2)":  addr_alt2,
        "Location Hint (Snippet)":      snippet_location,
        "Verification Confidence":      confidence,
        "Public Documents Found":       has_docs,
        # ── Hidden hyperlink data (excluded from displayed columns) ─────────
        "_li_url":  linkedin_url,
        "_fb_url":  fb_u,
        "_ig_url":  ig_u,
        "_tk_url":  tk_u,
        "_tw_url":  tw_u,
        "_gh_url":  gh_u,
        "_yt_url":  yt_u,
        "_rd_url":  rd_u,
        "_ot_url":  ot_u,
    }


# ---------------------------------------------------------------------------
# EXCEL STYLING
# ---------------------------------------------------------------------------

HEADER_FILL      = PatternFill("solid", fgColor="1F3864")
HEADER_FONT      = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
HYPERLINK_FONT   = Font(color="0563C1", underline="single", size=10)
NORMAL_FONT      = Font(size=10)


def style_sheet(ws) -> None:
    """Dark-blue header, auto column widths, frozen top row."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header = col_cells[0]
        header.fill      = HEADER_FILL
        header.font      = HEADER_FONT
        header.alignment = HEADER_ALIGN

        max_content = max(
            (len(str(c.value or "").split("\n")[0]) for c in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_content + 3, 14), 52
        )

    ws.freeze_panes = "A2"


def apply_hyperlinks(ws, df_sorted: pd.DataFrame, col_positions: dict) -> None:
    """
    Post-processes the worksheet to convert display-text cells into
    clickable hyperlinks using the _*_url hidden columns from df_sorted.
    """
    # Mapping: (display column name, hidden url column name)
    link_map = [
        ("LinkedIn",            "_li_url"),
        ("Facebook",            "_fb_url"),
        ("Instagram",           "_ig_url"),
        ("TikTok",              "_tk_url"),
        ("X / Twitter",         "_tw_url"),
        ("GitHub",              "_gh_url"),
        ("YouTube",             "_yt_url"),
        ("Reddit",              "_rd_url"),
        ("Other Social Platforms", "_ot_url"),
    ]

    for row_0idx, (_, row) in enumerate(df_sorted.iterrows()):
        excel_row = row_0idx + 2   # +1 for header row, +1 for 1-indexing

        for display_col, url_col in link_map:
            url     = row.get(url_col, "")
            display = row.get(display_col, "")
            if url and display and display_col in col_positions:
                cell           = ws.cell(row=excel_row, column=col_positions[display_col])
                cell.value     = display
                cell.hyperlink = str(url)
                cell.font      = HYPERLINK_FONT


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print(f"[*] Loading internal user registry from: {INPUT_EXCEL_USERS}")
    if not os.path.exists(INPUT_EXCEL_USERS):
        print(f"[!] File not found: {INPUT_EXCEL_USERS}")
        return

    try:
        df_base   = pd.read_excel(INPUT_EXCEL_USERS, engine="openpyxl")
        name_col  = df_base.columns[1]   # Column B — displayName
        dept_col  = df_base.columns[3]   # Column D — department
        title_col = df_base.columns[4]   # Column E — jobTitle
    except Exception as e:
        print(f"[!] Failed to parse spreadsheet: {e}")
        return

    user_registry: dict = {}
    for _, row in df_base.iterrows():
        emp = str(row.get(name_col, "")).strip()
        if emp and emp.lower() != "nan":
            user_registry[emp] = {
                "Internal Department": str(row.get(dept_col, "")),
                "Internal Job Title":  str(row.get(title_col, "")),
            }
    print(f"[*] Loaded {len(user_registry)} internal employee records.")

    print(f"[*] Loading OSINT JSON: {INPUT_RAW_JSON}")
    if not os.path.exists(INPUT_RAW_JSON):
        print(f"[!] File not found: {INPUT_RAW_JSON}")
        return

    with open(INPUT_RAW_JSON, "r", encoding="utf-8") as f:
        master_data = json.load(f)
    print(f"[*] Loaded {len(master_data)} employee OSINT records.")

    dossier_rows: list = []
    reference_rows: list = []

    print("[*] Running extraction pipeline...")
    total = len(master_data)

    for idx, employee in enumerate(master_data, 1):
        name    = employee.get("employee_name", "Unknown")
        assets  = employee.get("discovered_assets", [])
        profile = user_registry.get(name, {"Internal Department": "", "Internal Job Title": ""})

        print(f"    [{idx}/{total}] {name}")

        row = process_employee(employee, profile)
        dossier_rows.append(row)

        for asset in assets:
            if asset.get("url"):
                reference_rows.append({
                    "Employee Name":          name,
                    "Source URL":             asset["url"],
                    "Result Title":           asset.get("result_title", ""),
                    "Snippet":                asset.get("snippet", ""),
                    "Is Document":            "Yes" if asset.get("is_doc") else "No",
                    "Verification Confidence": row["Verification Confidence"],
                })

    # Build full DataFrame (includes _ columns)
    df_full = pd.DataFrame(dossier_rows)

    # Sort by confidence then name
    df_sorted = df_full.sort_values(
        by=["Verification Confidence", "Full Name"],
        ascending=[False, True],
    ).reset_index(drop=True)

    # Display columns only (exclude _ prefixed hidden URL columns)
    display_cols = [c for c in df_sorted.columns if not c.startswith("_")]
    df_display   = df_sorted[display_cols]

    # Column name → 1-indexed Excel column position
    col_positions = {col: idx + 1 for idx, col in enumerate(display_cols)}

    df_reference = pd.DataFrame(reference_rows)

    print(f"[*] Writing report: {OUTPUT_EXCEL_REPORT}")
    try:
        with pd.ExcelWriter(OUTPUT_EXCEL_REPORT, engine="openpyxl") as writer:

            # Sheet 1 — Main dossier (renamed to "Inside SEPTA")
            df_display.to_excel(writer, index=False, sheet_name="Inside SEPTA")

            # Sheet 2 — Raw link log
            df_reference.to_excel(writer, index=False, sheet_name="Source Links Log")

            # Style both sheets
            for sheet_name in ["Inside SEPTA", "Source Links Log"]:
                style_sheet(writer.sheets[sheet_name])

            # Apply hyperlinks to the main dossier sheet
            apply_hyperlinks(
                ws           = writer.sheets["Inside SEPTA"],
                df_sorted    = df_sorted,
                col_positions = col_positions,
            )

        print(f"[++] Report saved: {OUTPUT_EXCEL_REPORT}")

    except Exception as e:
        print(f"[!] Error writing Excel file: {e}")


if __name__ == "__main__":
    main()